from email import encoders
from email.header import Header
from email.mime.text import MIMEText
from email.utils import parseaddr, formataddr
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
import smtplib
from openpyxl import load_workbook,Workbook
import subprocess as sb
from openpyxl.styles import PatternFill, Alignment, Side, Border
import csv
import zipfile
import time
import os
import pandas as pd
from pandas import DataFrame,Series
#分组平均聚合
####数据可视化
##设置中文字体
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.pyplot import plot,savefig
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
#matplotlib.use('Agg')
##仿宋字体设置
plt.rcParams['font.family'] = ['FangSong']
paths ='./成绩分析清洗.csv'
target ='./学生个人成绩/'

def mkdir(path):
    folder = os.path.exists(path)
    if not folder:                 
        os.makedirs(path)            
    else:
        print("---  There is this folder!  ---")
##参数预置
tar_list =[]
test_list =[]
sums =0
##压缩文件模块
class zip:
    def get_zip(self,files,zip_name):
        zp=zipfile.ZipFile(zip_name,'w', zipfile.ZIP_DEFLATED)
        for file in files:
            zp.write(file)
        zp.close()
        time.sleep(1)

lest =os.listdir('./')
print(lest)
path1 =input("选择要打开的表格名称(输入序号): ")
path1 =lest[int(path1)-1]
path ='./'+path1
def open_f():
    lest =os.listdir('./')
    print(lest)
    path1 =input("选择要打开的表格名称(输入序号): ")
    path1 =lest[int(path1)-1]
    path ='./'+path1
    return path

def menu_all():
    print('''
    ********菜单********
    1.进入学生信息系统
    2.进入教师成绩系统
    0.退出
    ''')
def menu2():
    print('''
    ********成绩系统********
    1.学生成绩访问（查询，增加，删除，修改）
    2.学生成绩分析
    ''')
def menu2_add():
    print('''
    1.学生成绩查询
    2.学生成绩增加
    3.学生成绩删除
    4.学生成绩修改
    ''') 

def api_c():
    #调用c程序用以生成 xxx.txt
    sb.run(["info.exe"])

def visit():
    menu2_add()
    choice =input("请输入您的选择：")

    if(choice =='1'):#学生成绩查询
        wb =load_workbook(path)
        ws = wb.active
        name =input("请输入要查询的学生姓名: ")
        k =search(name)
        for cell in ws[2]:
            print("%s   "%cell.value,end ="")
        print("\n")
        for row in ws.iter_rows(min_row =k,max_row =k,min_col=1,max_col=6,values_only= True):
            print(row)
            
    elif(choice =='2'):#学生成绩增加
        print('''
            1.横向增加学生成绩
            2.纵向增加科目成绩
        ''')
        choices =input("请输入您的选择: ")
        if(choices =='1'):
            scores_x()
        elif(choices =='2'):
            scores_y()

    elif(choice =='3'):#学生成绩删除
        print('''
            1.横向删除学生的所有成绩信息
            2.纵向删除课程的所有成绩信息
        ''')
        choices =input("请输入您的选择: ")
        if(choices =='1'):
            del_X()
        elif(choices =='2'):
            del_y()
            
    elif(choice =='4'):#学生成绩修改
        name =input("请输入待修改学生姓名: ")
        lesson =input("请输入待修改科目: ")
        score =input("请输入待修改成绩数值: ")
        vary(name,lesson,score)

def search(name):
    k =1
    wb = load_workbook(path)
    ws = wb.active
    for cell in ws['C']:
        if(cell.value !=name):
            k =k+1
        else:
            break
    return k

def scores_x():
    wb =load_workbook(path)
    ws = wb.active
    adds =input("请依次输入以下信息以逗号间隔 例如 班级，学号，姓名，语文，数学，英语 :")
    lists =adds.split(",")
    ws.append(lists)
    rows =ws.max_row
    for cell in ws[rows]:
        cell.alignment = align
    wb.save(path)

def scores_y():
    wb =load_workbook(path)
    ws = wb.active
    cols =ws.max_column
    cols =['A','B','C','D','E','F','G','H'][cols-1]
    rows =ws.max_row
    lesson =input("输入要增加的科目:")
    ws[cols+'2'] =lesson
    for i in range(3,rows+1):
        print(ws['C'+str(i)].value)
        score =input("输入该同学"+lesson+'成绩: ' )
        ws[cols+str(i)] =score
        ws[cols+str(i)].alignment = align
        print("\n")
    wb.save(path)

def del_X():
    wb =load_workbook(path)
    ws = wb.active
    adds =input("请输入要删除的学生姓名: ")
    k =search(adds)
    ws.delete_rows(k) #删除从第一行开始算的2行内容
    wb.save(path)

def del_y():
    wb =load_workbook(path)
    ws = wb.active
    adds =input("请输入要删除的课程名称: ")
    k =4
    wb =load_workbook(path)
    ws = wb.active
    for i in ['D','E','F','G','H']:
        if(ws[i+str(2)].value !=adds):
            k =k+1
        else:
            break
    ws.delete_cols(k) #删除从第一列开始算的2列内容
    wb.save(path)

def vary(name,lesson,score):
    rows =search(name)
    wb =load_workbook(path)
    ws = wb.active
    k =4
    for i in ['D','E','F','G','H']:
        if(ws[i+str(2)].value !=lesson):
            k =k+1
        else:
            break
    cols =['A','B','C','D','E','F','G','H'][k-1]
    ws[cols+str(rows)] =score
    ws[cols+str(rows)].alignment = align
    wb.save(path)



def analyze_fun(ws1,ws2):#成绩分析
    fo =open('期中期末成绩汇总表.csv','w',encoding= 'utf-8')
    lists =[]
    lists.append(['班级','姓名','期中语文','期中数学','期中英语','期末语文','期末数学','期末英语','\n'])
    lists =",".join(lists[0])
    fo.write(lists)
    str1 =''
    lists =[]
    for row in ws1.iter_rows(min_row =3,max_row =782,min_col =2,max_col =6,values_only =True):
        for i in row:
            if i ==row[-1]:
                str1 =str1+str(i)
            else:
                str1 =str1+str(i)+','
        lists.append(str1)
        str1 =''
    x =0
    for row in ws2.iter_rows(min_row =3,max_row =782,min_col =4,max_col =6,values_only =True):
        for i in row:
            if i ==row[-1]:
                lists[x] =lists[x]+','+str(i)+'\n'
            else:
                lists[x] =lists[x]+','+str(i)
        x =x+1
    for i in lists:
        fo.write(i)
    fo.close()

def analyze_clean():
    read_csv =pd.read_csv('./期中期末成绩汇总表.csv',encoding ='utf-8')
    read_csv =read_csv.drop('Unnamed: 8',axis=1)
    #查找缺失值
    read_csv.isna()
    #删除缺失值
    read_csv =read_csv.dropna()
    #can't del repetitive value
    analy =read_csv.describe()
    analy.to_csv('./成绩描述性统计数据.csv',encoding ='utf-8')
    #描述性统计分析后没有异常值存在，为简化不必清洗异常值
    #清洗后数据存储
    read_csv.to_csv('./成绩分析清洗.csv',encoding ='utf-8')

def analyze_class():
    grade_df =pd.read_csv('./成绩分析清洗.csv',encoding ='utf-8')
    grade_df
    #平均成绩聚合
    grade_mid_chinese =grade_df.groupby('班级')['期中语文'].mean()
    grade_last_chinese =grade_df.groupby('班级')['期末语文'].mean()
    grade_mid_math =grade_df.groupby('班级')['期中数学'].mean()
    grade_last_math =grade_df.groupby('班级')['期末数学'].mean()
    grade_mid_eng =grade_df.groupby('班级')['期中英语'].mean()
    grade_last_eng =grade_df.groupby('班级')['期末英语'].mean()
    grade_df2 =DataFrame({'期中语文':grade_mid_chinese,'期末语文':grade_last_chinese,'期中数学':grade_mid_math,'期末数学':grade_last_math,'期中英语':grade_mid_eng,'期末英语':grade_last_eng})
    grade_df2.to_csv('./班级分组平均成绩聚合.csv',encoding ='utf-8')
    #成绩标差聚合
    grades_mid_chinese =grade_df.groupby('班级')['期中语文'].std()
    grades_last_chinese =grade_df.groupby('班级')['期末语文'].std()
    grades_mid_math =grade_df.groupby('班级')['期中数学'].std()
    grades_last_math =grade_df.groupby('班级')['期末数学'].std()
    grades_mid_eng =grade_df.groupby('班级')['期中英语'].std()
    grades_last_eng =grade_df.groupby('班级')['期末英语'].std()
    grades_df2 =DataFrame({'期中语文':grades_mid_chinese,'期末语文':grades_last_chinese,'期中数学':grades_mid_math,'期末数学':grades_last_math,'期中英语':grades_mid_eng,'期末英语':grades_last_eng})
    grades_df2.to_csv('./班级分组标差成绩聚合.csv',encoding ='utf-8')
    
    #画布生成
    plt.figure(figsize =(15,15))
    fun_read =pd.read_csv('./班级分组平均成绩聚合.csv',encoding ='utf-8')
    xclass =fun_read['班级']
    ychinese =fun_read[['期中语文','期末语文']]
    plt.plot(xclass,ychinese,linewidth =3,marker ='o',markersize =10,markerfacecolor ='w')
    #图表标题及字体大小
    plt.title('各班语文平均成绩折线图',fontsize =20)
    #坐标轴刻度字体
    plt.xticks(fontsize =15,rotation =90)
    plt.yticks(fontsize =15)
    plt.xlabel('班级',fontsize =15)
    plt.ylabel('成绩（分）',fontsize =15)
    plt.legend(['期中语文','期末语文'])
    savefig('./各班语文平均成绩折线图.png')
    plt.close()

    
    funs_read =pd.read_csv('./班级分组标差成绩聚合.csv',encoding ='utf-8')
    plt.figure(figsize =(15,15))
    ####柱状图基本设计
    # 设置 x/y 坐标值
    x =funs_read['班级']
    y =funs_read['期中语文']
    plt.plot(x, y, color='dodgerblue')
    plt.title('年级班级语文成绩标差分布',fontdict ={
            'family': 'FangSong', 'color': 'black', 'weight': 'bold', 'size': 25})
    plt.xticks(fontsize=18,rotation =90)
    plt.yticks(fontsize=12)
    plt.xlabel('班级', fontsize=15)
    plt.ylabel('标差稳定性', fontsize=17)
    plt.bar(x, height=y, color='darkorange',width=0.6,alpha=0.6)
    plt.legend(['稳定性变化','稳定性分布'])
    savefig('./各班语文标差成绩图.png')
    plt.close()
    #for a,b in zip(x,y):
        #plt.text(a, b, b, ha='center', va='bottom', fontsize=12)
def analy_stu(paths):
    sums =0
    with open(paths,'r',encoding ='utf-8',newline ="") as csv_file:
        csvs_reader =csv.DictReader(csv_file)
        headers =csvs_reader.fieldnames
        for row in csvs_reader:
            target_file =row['姓名']+'.csv'
            fil =target+row['姓名']+'/'
            mkdir(fil)
            test_list.append(fil)
            target_file =fil+target_file
            tar_list.append(target_file)
            sums =sums+1
            with open(target_file,'w',encoding ='utf-8',newline ="") as csv_f:
                csv_w =csv.DictWriter(csv_f,headers)
                csv_w.writeheader()
                csv_w.writerow(row)
    i=0
    for tar_file in tar_list[0:1]:###########
        f =pd.read_csv(tar_file,encoding ='utf-8')
        plt.figure(figsize =(15,15))
        xclass =f['姓名']
        ychinese =f[['期中语文','期末语文']]
        plt.plot(xclass,ychinese,linewidth =3,marker ='o',markersize =10,markerfacecolor ='w')
        #图表标题及字体大小
        plt.title('语文成绩折线图',fontsize =20)
        #坐标轴刻度字体
        plt.xticks(fontsize =15)
        plt.yticks(fontsize =15)
        plt.xlabel('姓名',fontsize =15)
        plt.ylabel('成绩（分）',fontsize =15)
        plt.legend(['期中语文','期末语文'])
        savefig(test_list[i]+'语文成绩折线图.png')
        i =i+1
        plt.close()
        ##压缩模块
    i =0
    name =[]

    lists =[]
    lest =[]
    for fo in test_list[0:1]:#########
        lists =os.listdir(fo)
        for x in lists:
            path1 =fo+lists[0]
            lest.append(path1)
            path2 =fo+lists[1]
            lest.append(path2)
        z =zip()
        zip_file =fo+'成绩.zip'
        name.append(zip_file)
        z.get_zip(lest,zip_file)
        time.sleep(2)
        i =i+1
        sums =sums-1
        lest =[]
        lists =[]
        print("{}个文件已完成，剩余{}个预计需要{}分钟".format(i,sums,sums*3/60))

def maile(aim_account):
    account = input('请输入邮箱账户：')
    token = input('请输入邮箱授权码：')
    # 设置邮箱服务器，端口
    smtp = smtplib.SMTP_SSL('smtp.qq.com', 465)
    # 登录qq邮箱
    smtp.login(account, token)
    content ='本学期成绩已整理完成，现在对你的成绩单独发送'
    content =content+'详情见附件内容'
    email_content = MIMEText(content, 'plain', 'utf-8')
#for tar in name:
    #passhttp://localhost:8888/notebooks/Untitled3.ipynb#
    tar ='./学生个人成绩/高健玮/成绩.zip'#########
    f =open(tar,'rb')
    # 设置附件的MIME和文件名，这里是rar类型:
    fil = MIMEBase('zip', 'zip', filename='成绩单.zip')
    # 加上必要的头信息:
    fil.add_header('Content-Disposition', 'attachment', filename='成绩单')
    fil.add_header('Content-ID', '<0>')
    fil.add_header('X-Attachment-Id', '0')
    # 把附件的内容读进来:
    fil.set_payload(f.read())
    # 用Base64编码
    encoders.encode_base64(fil)
    #添加到MIMEMultipart
    msg = MIMEMultipart()
    msg.attach(fil)
    f.close()
    msg.attach(email_content)
    # 设置发送者信息
    msg['From'] = '贾'
    msg['To'] = '各位同事们' 
    msg['Subject'] = '测试'
    # 发送邮件
    smtp.sendmail(account, aim_account, msg.as_string())  
    # 关闭邮箱服务
    smtp.quit() 

choice =5###学生成绩管理后期用openpyxl改进
align = Alignment(horizontal='right')
while(choice !='0'):
    menu_all()
    choice =input("请输入您的选择:")
    
    if(choice =='2'):#教师端操作
        menu2()
        choice =input('请输入您的选择：')

        if(choice =='1'):#学生成绩访问
            visit()
        
        elif(choice =='2'):#学生成绩分析
            print('''
            1.班级总体分析
            2.单位学生分析
            ''')
            choice =input("请输入你的选择: ")
            if(choice =='1'):#总体分析
                print("请打开两个xlsx表格")
                #打开 期中、期末两个xlsx表格
                path1 =open_f()
                wb1 =load_workbook(path1)
                ws1 =wb1.active
                path2 =open_f()
                wb2 =load_workbook(path2)
                ws2 =wb2.active
                #用函数封装具体分析
                analyze_fun(ws1,ws2)
                analyze_clean()
                analyze_class()
                
            
            elif(choice =='2'):#单位学生分析
                analy_stu(paths)
                aim_account =input('请输入要发送的目标邮箱: ')
                maile(aim_account)
    
    elif(choice =='1'):
        api_c()



    menu_all()
    choice =input("请输入您的选择:")       
